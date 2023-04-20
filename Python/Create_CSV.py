import csv
import openpyxl
import pandas as pd
import xlrd


def create_new_countries_csv():
     
    country_names_list = [] # H lista tou countries.csv.
    country_names_area_list = []    # H lista tou country_names_area.
    
    with open('countries.csv') as file:
        file_data = csv.reader(file, delimiter = ',')  # load csv file separated by commas.
        country_names_list.append(next(file_data))
        country_names_list[-1].append('Country_area')   # Kane append allo ena header sthn teleutaia thesi tou 1ou row.
        for row in file_data:
            country_names_list.append(row)
    for row in country_names_list:
        for column in range(23):
            if row[column] == '':
                row[column] = 'NULL'    # Opou uparxei keno sto csv bainei NULL.
                if row[10].startswith('+'):
                    row[10] = 'NULL'
    
    with open('country_names_area.csv') as file:
        file_data = csv.reader(file, delimiter = ',')  # load csv file separated by commas.
        for row in file_data:
            country_names_area_list.append(row)

    # Diplh for oste na brisko ama uparxei h xora tou country_names_area.csv sto countries.csv:

    for x in range(len(country_names_area_list)):
        for y in range(len(country_names_list)):
            if country_names_area_list[x][1] == country_names_list[y][4]:
                country_names_list[y].append(country_names_area_list[x][2]) # An uparxei, tote sto row pou to brhka prostheto sto column tou neou Header Country_area ton kodiko tou country_names_area.csv.
    # Merikes xores tou countries.csv den uparxoun sto country_names_area.csv opote sthn thesi twn kenwn pou dhmiourgounte vazo NULL:

    for x in range(len(country_names_list)):
        if len(country_names_list[x]) == 24:
            country_names_list[x].append('NULL')

    # Dhmiourgo ena kainourio csv arxeio kai tou fortono thn country_names_list:

    with open('countries_output.csv', 'w') as file:
        writer = csv.writer(file, lineterminator = '\n')

        for row in range(len(country_names_list)):
            writer.writerow(country_names_list[row])

    return country_names_list


def sorting_country_year(filename):

    # Kano sort me vash to onoma ths xoras kai th xronia:

    header = []

    with open(filename, 'r') as csvfile:
        reader = csv.reader(csvfile)

        for i, row in enumerate(reader):    # Ksexorizo to header prin kano sort.
            if i == 0:
                header.append(row)
                break

        rows = list(reader)
        rows.sort(key=lambda row: (row[1], row[2])) # Kano sort me vash to country_name kai ta year.

    # Dhmiourgo ena kainourio csv arxeio kai tou fortono to sorted list rows:

    filename = filename.replace('.csv', '_')
    with open(f'{filename}output.csv', 'w', newline='') as csvfile:
        writer = csv.writer(csvfile)    
        writer.writerows(header)    # Grafo prota to header.
        writer.writerows(rows)      # Sumplhrono ta sorted dedomena.

    return 1


def add_iso_code(countries_list, csvfile):
    
    csv_list = []
    no_valid_name = []
    flag = 0
    count = 0
    index = []

    with open(csvfile) as file:
        file_data = csv.reader(file, delimiter = ',')  # load csv file separated by commas.
        for row in file_data:
            csv_list.append(row)

    csv_list[0].append('ISO_Code')   # Kane append allo ena header sthn teleutaia thesi tou 1ou row.

    # Kano append ta ISO_CODE:

    for x in range(len(csv_list) - 1):
        for y in range(len(countries_list) - 1):
            if csv_list[x + 1][1] == countries_list[y + 1][4]:
                csv_list[x + 1].append(countries_list[y + 1][2])

    # Gemizo thn no_valid_name me tis xores pou den uparxoun sto countries_list:

    for x in range(len(csv_list) - 1):
        for y in range(len(countries_list) - 1):
            if csv_list[x + 1][1] == countries_list[y + 1][4]:
                flag = 1
                break
        if flag == 0:
            if count == 0:
                no_valid_name.append(csv_list[x + 1][1])
                count += 1
            else:
                if no_valid_name[-1] != csv_list[x + 1][1]:
                    no_valid_name.append(csv_list[x + 1][1])
        flag = 0

    # Diagrafo tis grammes pou oi xores tous den anhkoun sto countries:

    for row in range(len(csv_list) - 1):
        for x in no_valid_name:
            if csv_list[row + 1][1] == x:
                index.append(row + 1)
    
    index.sort(reverse=True)

    for x in index:
        del csv_list[x]

    #csvfile = csvfile.replace('_.csv', '_.csv')
    with open(f'{csvfile}', 'w', newline='') as csvfile:
        writer = csv.writer(csvfile)    
        writer.writerows(csv_list)

    return 1


########################################################################################################################

def create_csv_from_xlsx(filename):

    # Load the XLSX file
    workbook = openpyxl.load_workbook(filename)

    # Iterate over each sheet in the workbook
    for sheet_name in workbook.sheetnames:
        # Create a new workbook for each sheet
        new_workbook = openpyxl.Workbook()
        new_sheet = new_workbook.active

        # Copy the values from the original sheet to the new sheet
        sheet = workbook[sheet_name]
        for row in sheet:
            for cell in row:
                new_sheet[cell.coordinate].value = cell.value

        # Save the new workbook with the sheet's name as the filename
        new_workbook.save(f'{sheet_name}.xlsx')
        
    return 1


def change_xlsx_to_csv(filename):

    # Load the XLSX file
    df = pd.read_excel(filename)

    # Save the dataframe as a CSV file
    filename = filename.replace('.xlsx', '_OUT')
    df.to_csv(F'{filename}.csv', index=False)

    return 1


def fiveyear_to_oneyear(filename):

    csv_list = []   # Ths fortono to arxeio csv.
    count = 0   # Metrhths gia tis Xores.
    count1 = 0  # Metrhths gia ta Xronia.
    count2 = 1  # Metrhths gia tis Metrhseis.

    with open(filename) as file:
        file_data = csv.reader(file, delimiter = ',')  # load csv file separated by commas.
        for row in file_data:
            csv_list.append(row)

    my_list = [[] for i in range(29*(len(csv_list) - 1) + 1)]   # ftiaxno mia adeia 2d list me (29*(len(csv_list) - 1) + 1) rows.

    my_list[0].insert(0, csv_list[0][0])    # Insert to "Country" sto [0,0].

    # Vazo oles tis xores sthn proth sthlh, 29 fores thn kathemia:

    for x in range(len(csv_list) - 1):
        for y in range(29):
            my_list[count + 1].insert(0, csv_list[x + 1][0])
            count += 1

    my_list[0].append('Year')   # Append to "Year" sto [0,1].

    # Antistoixo oles tis xronies 1990-2018 se kathe xora:

    for row in range(len(my_list) - 1):
        my_list[row + 1].append(1990 + count1)
        count1 += 1
        if count1 == 29:
            count1 = 0   

    my_list[0].append(filename.replace('_OUT.csv', '')) # Append to onoma ths metrhshs sto [0,2].

    # Vazo ola tis metrhseis tou csv_list sto my_list.

    for row in range(len(csv_list) - 1):
        for col in range(4):
            for x in range(5):
                my_list[count2].append(csv_list[row + 1][col + 1])
                count2 += 1
        for y in range(9):
            my_list[count2].append(csv_list[row + 1][5 + y])
            count2 += 1
    
    # Elegxo gia '' kai '..', opou ta vrisko vazo 'NULL':

    for row in range(len(my_list)):
        if my_list[row][2] == '..' or my_list[row][2] == '':
            my_list[row][2] = 'NULL'

    filename = filename.replace('.csv', '')
    with open(f'{filename}.csv', 'w', newline='') as csvfile:
        writer = csv.writer(csvfile)    
        writer.writerows(my_list)

    return 1


def create_years_column(filename):

    my_list = [[] for i in range(29*208 + 1)] # ftiaxno mia adeia 2d list me (29*208 + 1) rows gt 29 einai ta xronia kai 208 oi xores sta Income Index_OUT kai GNI per capita_OUT. 
    csv_list = []   # Ths fortono to arxeio csv.
    count = 0   # Metrhths gia tis Xores.
    count1 = 0  # Metrhths gia ta Xronia.
    count2 = 1  # Metrhths gia tis Metrhseis.

    with open(filename) as file:
        file_data = csv.reader(file, delimiter = ',')  # load csv file separated by commas.
        for row in file_data:
            csv_list.append(row)

    my_list[0].insert(0, csv_list[0][0])    # Insert to "Country" sto [0,0].

    # Vazo oles tis xores sthn proth sthlh, 29 fores thn kathemia:

    for x in range(len(csv_list) - 1):
        for y in range(29):
            my_list[count + 1].insert(0, csv_list[x + 1][0])
            count += 1

    my_list[0].append('Year')   # Append to "Year" sto [0,1].

    # Vazo oles tis xronies 1990-2018 se kathe xora:

    for row in range(len(my_list) - 1):
        my_list[row + 1].append(1990 + count1)
        count1 += 1
        if count1 == 29:
            count1 = 0

    my_list[0].append(filename.replace('_OUT.csv', '')) # Append to onoma ths metrhshs sto [0,2].

    # Vazo ola tis metrhseis tou csv_list sto my_list.

    for row in range(len(csv_list) - 1):
        for col in range(29):
            my_list[count2].append(csv_list[row + 1][col + 1])
            count2 += 1

    # Elegxo gia '' kai '..', opou ta vrisko vazo 'NULL':

    for row in range(len(my_list)):
        if my_list[row][2] == '..' or my_list[row][2] == '':
            my_list[row][2] = 'NULL'

    filename = filename.replace('.csv', '')
    with open(f'{filename}.csv', 'w', newline='') as csvfile:
        writer = csv.writer(csvfile)    
        writer.writerows(my_list)

    return 1


def estimated_GNI_female_male(filename):

    csv_list = []   # Ths fortono to arxeio csv.
    count = 0   # Metrhths gia tis Xores.
    count1 = 0  # Metrhths gia ta Xronia.
    count2 = 1  # Metrhths gia tis Metrhseis.

    with open(filename) as file:
        file_data = csv.reader(file, delimiter = ',')  # load csv file separated by commas.
        for row in file_data:
            csv_list.append(row)

    my_list = [[] for i in range(24*(len(csv_list) - 1) + 1)]   # ftiaxno mia adeia 2d list me (24*(len(csv_list) - 1) + 1) rows.

    my_list[0].insert(0, csv_list[0][0])    # Insert to "Country" sto [0,0].

    # Vazo oles tis xores sthn proth sthlh, 24 fores thn kathemia:

    for x in range(len(csv_list) - 1):
        for y in range(24):
            my_list[count + 1].insert(0, csv_list[x + 1][0])
            count += 1

    my_list[0].append('Year')   # Append to "Year" sto [0,1].

    # Antistoixo oles tis xronies 1990-2018 se kathe xora:

    for row in range(len(my_list) - 1):
        my_list[row + 1].append(1995 + count1)
        count1 += 1
        if count1 == 24:
            count1 = 0   

    my_list[0].append(filename.replace('_OUT.csv', '')) # Append to onoma ths metrhshs sto [0,2].

    # Vazo ola tis metrhseis tou csv_list sto my_list.

    for row in range(len(csv_list) - 1):
        for col in range(3):
            for x in range(5):
                my_list[count2].append(csv_list[row + 1][col + 1])
                count2 += 1
        for y in range(9):
            my_list[count2].append(csv_list[row + 1][4 + y])
            count2 += 1
    
    # Elegxo gia '' kai '..', opou ta vrisko vazo 'NULL':

    for row in range(len(my_list)):
        if my_list[row][2] == '..' or my_list[row][2] == '':
            my_list[row][2] = 'NULL'

    filename = filename.replace('.csv', '')
    with open(f'{filename}.csv', 'w', newline='') as csvfile:
        writer = csv.writer(csvfile)    
        writer.writerows(my_list)

    return 1


def labour_share_GDP_CSV(filename):

    csv_list = []   # Ths fortono to arxeio csv.
    count = 0   # Metrhths gia tis Xores.
    count1 = 0  # Metrhths gia ta Xronia.
    count2 = 1  # Metrhths gia tis Metrhseis.

    with open(filename) as file:
        file_data = csv.reader(file, delimiter = ',')  # load csv file separated by commas.
        for row in file_data:
            csv_list.append(row)

    my_list = [[] for i in range(19*(len(csv_list) - 1) + 1)]   # ftiaxno mia adeia 2d list me (19*(len(csv_list) - 1) + 1) rows.

    my_list[0].insert(0, csv_list[0][0])    # Insert to "Country" sto [0,0].

    # Vazo oles tis xores sthn proth sthlh, 29 fores thn kathemia:

    for x in range(len(csv_list) - 1):
        for y in range(19):
            my_list[count + 1].insert(0, csv_list[x + 1][0])
            count += 1

    my_list[0].append('Year')   # Append to "Year" sto [0,1].

    # Antistoixo oles tis xronies 1990-2018 se kathe xora:

    for row in range(len(my_list) - 1):
        my_list[row + 1].append(2000 + count1)
        count1 += 1
        if count1 == 19:
            count1 = 0   

    my_list[0].append(filename.replace('_OUT.csv', '')) # Append to onoma ths metrhshs sto [0,2].

    # Vazo ola tis metrhseis tou csv_list sto my_list.

    for row in range(len(csv_list) - 1):
        for col in range(2):
            for x in range(5):
                my_list[count2].append(csv_list[row + 1][col + 1])
                count2 += 1
        for y in range(9):
            my_list[count2].append(csv_list[row + 1][3 + y])
            count2 += 1
    
    # Elegxo gia '' kai '..', opou ta vrisko vazo 'NULL':

    for row in range(len(my_list)):
        if my_list[row][2] == '..' or my_list[row][2] == '':
            my_list[row][2] = 'NULL'

    filename = filename.replace('.csv', '')
    with open(f'{filename}.csv', 'w', newline='') as csvfile:
        writer = csv.writer(csvfile)    
        writer.writerows(my_list)

    return 1

    

#########################################################################################################################################################

def main():

    my_countries =  create_new_countries_csv()

    sorting_country_year('age_specific_fertility_rates.csv')
    add_iso_code(my_countries, 'age_specific_fertility_rates_output.csv')

    sorting_country_year('birth_death_growth_rates.csv')
    add_iso_code(my_countries, 'birth_death_growth_rates_output.csv')

    sorting_country_year('midyear_population_age_sex.csv')
    add_iso_code(my_countries, 'midyear_population_age_sex_output.csv')

    sorting_country_year('midyear_population_5yr_age_sex.csv')
    add_iso_code(my_countries, 'midyear_population_5yr_age_sex_output.csv')

    sorting_country_year('mortality_life_expectancy.csv')
    add_iso_code(my_countries, 'mortality_life_expectancy_output.csv')

    sorting_country_year('midyear_population.csv')
    add_iso_code(my_countries, 'midyear_population_output.csv')

    ##################################################################################################    

    create_csv_from_xlsx('Income by Country.xlsx')

    change_xlsx_to_csv('Domestic credits.xlsx')
    change_xlsx_to_csv('Estimated GNI female.xlsx')
    change_xlsx_to_csv('Estimated GNI male.xlsx')
    change_xlsx_to_csv('GDP per capita.xlsx')
    change_xlsx_to_csv('GDP total.xlsx')
    change_xlsx_to_csv('GNI per capita.xlsx')
    change_xlsx_to_csv('Gross fixed capital formation.xlsx')
    change_xlsx_to_csv('Income Index.xlsx')
    change_xlsx_to_csv('Labour share of GDP.xlsx')

    create_years_column('Income Index_OUT.csv')
    create_years_column('GNI per capita_OUT.csv')
    fiveyear_to_oneyear('Domestic credits_OUT.csv')
    fiveyear_to_oneyear('GDP per capita_OUT.csv')
    fiveyear_to_oneyear('GDP total_OUT.csv')
    fiveyear_to_oneyear('Gross fixed capital formation_OUT.csv')
    estimated_GNI_female_male('Estimated GNI female_OUT.csv')
    estimated_GNI_female_male('Estimated GNI male_OUT.csv')
    labour_share_GDP_CSV('Labour share of GDP_OUT.csv')

    return 1


###########################################################################################
#                                          START                                          #
###########################################################################################

main()
